import random
import folium
import openpyxl as opx


# main file
file = opx.open(
    "main1.xlsx")  # open excel file
base_station = {}  # Создаем словарь

sheet = file.active  # Помечам лист файла
file_score = len(sheet["C"]) - 1  # Задаем длинну столбца с которым работаем
print(file_score)  # проверяем длину
row = 2  # задаем номер строки, с которой будем работать
colum = 5  # Задаем номер столбца, с которым будем работать на начальном этапе
text = int(input("Номер станции - "))  # Вводим номер станции
base_station[text] = 0  # Присваиваем словарю значение к номеру станции, который ввели выше

tenRow = 2  # Опять вводим номер строки для удобства работы, чтобы не сверять номера с другими и не править в дальнейшем
sheet.cell(row=tenRow,
           column=13).value = text  # присваиваем строке tenRow в колонке №13 номер станции, который ввели выше
while text != 0:  # цикл, который будет добавлять ноемера стаций в таблицу, для последуюющей работы
    text = int(input("Номер станции - "))
    if text != 0:
        base_station[text] = 0
        sheet.cell(row=tenRow + 1, column=13).value = text  # Присваивани строке в Excel номера станции
        tenRow += 1
    else:
        continue
score = len(base_station)  # получаем длину словаря

create_random = random.randint(0, score - 1)

rand = random.randint(1, score)  # создаем переменную для процесса рандома
base_station_list = list(base_station.keys())
base_station_r = random.choice(base_station_list)  # присваиваем рандомное число переменной

for i in range(file_score):  # создаем цикл равный длинне столбца с номерами, для проставления номеров станций
    sheet.cell(row=row, column=colum).value = base_station_list[create_random]
    create_random = random.randint(0, score - 1)

    sheet.cell(row=row, column=colum + 1).value = str(
        base_station.keys())  # с ее помощью можно узнать количество всех участвующих в процессе станций

    print("номера:", sheet["C" + str(row)].value, "Количество станций:",
          sheet["D" + str(row)].value, "Номер станции:", sheet["D" + str(row)].value,
          sep="\n")  # выводим данные для проверки

    base_station[base_station_r] += 1
    base_station_r = random.choice(base_station_list)
    rand = random.randint(1, score)  # меняем рандомное число
    row += 1
file.save("main1.xlsx")  # сохраняем файл

# сортировка словоря

sort_BS = sorted(base_station.items(), key=lambda x: x[1])  # сортируем данные, создавая рейтинг от меньшего к боьшему
print('Рейтинг, но только в словаре: ', dict(sort_BS), end="\n")  # выводим данные для проверки

# создаем базу данных координат
# Excel База данных
coordE = opx.open('Coordinates.xlsx')
coorExcel = coordE.active
print('# for by 13-14 column')  # выводим для обозначения следующего этапа
rowAdd = 2  # номер строки
columAdd = 5  # номер столбца
elevenColumn = 14  # номер столбца, и не важно как он назван
baseCount = 1  # переменная для проверки номеров


for i in range(len(base_station)):

    print(sheet.cell(row=rowAdd, column=elevenColumn - 1).value, '-',
          baseCount)  # сверяем номер для проверки и то что записано в таблице
    if sheet.cell(row=rowAdd,
                  column=elevenColumn - 1).value == baseCount:  # создаем вопрос для проверки и добавления координат
        # к соответствующему номеру
        print('baseCount =', baseCount)
        sheet.cell(row=rowAdd, column=elevenColumn).value = coorExcel.cell(row=baseCount + 1,
                                                                           column=1).value  # присваиваем координаты
        print(sheet.cell(row=rowAdd, column=elevenColumn).value, coorExcel.cell(row=baseCount + 1,
                                                                                column=1).value)
        sheet.cell(row=rowAdd, column=elevenColumn + 1).value = coorExcel.cell(row=baseCount + 1,
                                                                               column=2).value  # присваиваем координаты
        print(baseCount, rowAdd,
              sheet.cell(row=rowAdd, column=elevenColumn).value,
              coorExcel.cell(row=baseCount + 1,
                             column=1).value, sep='\n')  # выводим данные для проверки
    baseCount += 1
    rowAdd += 1
file.save("main1.xlsx")
coordE.close()

rowAdd1 = 2  # номер строки
columAdd1 = 5  # номер столбца
columCoordinates1 = 7  # номер столбца

print('Номера', sheet.cell(row=4, column=3).value, 'Станции', sheet.cell(row=4, column=5).value,
      '1 координата', sheet.cell(row=4, column=7).value, '2 координата',
      sheet.cell(row=4, column=8).value)  # выводим для сверки
maxStation = max(base_station)  # максимальное число из словаря
print('max(base_station) =', maxStation)  # выводим
sheet.cell(row=1, column=13).value = maxStation  # добавляем в таблицу, для проверки в дальнейшем
print(sheet.cell(row=1, column=13).value)  # выводим что находится в этой строке

file.close()  # закрываем файл

# создаем блок с картой, есть мысль потом его  в функцию закрыть, чтобы много места не занимал
print('base_station -', base_station)  # для сверки выводим словарь
print('sort_BS', sort_BS)  # а также рейтинг
url = opx.open(
    "main1.xlsx")  # открываем файл excel

sheet = url.active  # объявляем лист с которым будем работать

# <color>
print(sort_BS)
print(len(sort_BS))
for i in range(len(sort_BS)):  # вписываем количество телефонных номеров у станции, в строчку к этой станцие
    text = False
    fuondNum = 0
    print(fuondNum, '- fuondNum')
    print(i, '- i')
    print(sort_BS[i][0], 'sort_BS[i][1] -', sort_BS[fuondNum][1])
    while text != True:
        if sheet.cell(row=i + 2, column=13).value == sort_BS[fuondNum][0]:
            print(sheet.cell(row=i + 2, column=13).value,
                  sort_BS[fuondNum][0])
            sheet.cell(row=i + 2, column=16).value = sort_BS[fuondNum][1]
            text = True
        else:
            fuondNum += 1
            print(fuondNum, '- fuondNum + 1')
url.save("main1.xlsx")

colorFile = opx.open('color_base.xlsx')
count = 0
rowLen = 2
colorList = colorFile.active

# ----------ЦИКЛ ПРОЦЕНТ-----------------
fileCount = file_score // len(sort_BS)
c = 2


while c != len(sort_BS):
    someNum = sheet.cell(row=rowLen,
                         column=16).value  # Выделяем строчку и столбец, в которой записано кол-во присвоенных номеров
    sl = []
    phoneN = sheet[f"C{c}"].value
    print(phoneN)
    for i in str(phoneN):
        sl.append(int(i))
    slSum = sum(sl[-4:])
    print(sl[-4:])
    print(slSum, " = slSum")
    c += 1
    print(c, " = c")
    print(len(sheet["C"]) + 1)

    percent1 = range(0, 4)
    percent2 = range(4, 7)
    percent3 = range(7, 10)
    percent4 = range(10, 14)
    percent5 = range(14, 17)
    percent6 = range(17, 21)
    percent7 = range(21, 25)
    percent8 = range(25, 28)
    percent9 = range(28, 33)
    percent10 = range(33, 36)

    # цикл на подбор числа к диапазону уменьшенных показателей
    if slSum in percent1:  # проверяем наличие числа в диапазонеот среднего числа, до ближайшего %
        print('range =', percent1)  # выводим для проверки диапазона
        print(slSum, '= slSum')  # выводим для проверки числа в строке Excel
        colorLen = 5  # номер строки в базе данных цветов, с нужным цветом
        count += 1  # сверка конечного количества выполненых циклов к количеству станций (1 к 1)
        sheet.cell(row=rowLen,
                   column=17).value = \
            colorList.cell(row=colorLen,
                           column=2).value  # присваивание цветового номера к номеру БС в Excel
        sheet.cell(row=rowLen,
                   column=18).value = colorLen  # присваивание срочного номера табл. цветов, к номеру БС в табл Номеров
        # дальше будут такие же компоненты
        print(colorList.cell(row=colorLen, column=1).value)
    elif slSum in percent2:
        print('range =', percent2)
        print(slSum, '= slSum')
        colorLen = 4
        count += 1
        sheet.cell(row=rowLen, column=17).value = colorList.cell(row=colorLen, column=2).value
        sheet.cell(row=rowLen, column=18).value = colorLen
        print(colorList.cell(row=colorLen, column=1).value)
    elif slSum in percent3:
        print('range =', percent3)
        print(slSum, '= slSum')
        colorLen = 3
        count += 1
        sheet.cell(row=rowLen, column=17).value = colorList.cell(row=colorLen, column=2).value
        sheet.cell(row=rowLen, column=18).value = colorLen
        print(colorList.cell(row=colorLen, column=1).value)
    elif slSum in percent4:
        print('range =', percent4)
        print(slSum, '= slSum')
        colorLen = 2
        count += 1
        sheet.cell(row=rowLen, column=17).value = colorList.cell(row=colorLen, column=2).value
        sheet.cell(row=rowLen, column=18).value = colorLen
        print(colorList.cell(row=colorLen, column=1).value)
    elif slSum in percent5:
        print('range =', percent5)
        print(slSum, '= slSum')
        colorLen = 1
        count += 1
        sheet.cell(row=rowLen, column=17).value = colorList.cell(row=colorLen, column=2).value
        sheet.cell(row=rowLen, column=18).value = colorLen
        print(colorList.cell(row=colorLen, column=1).value)
    # ----------------PLUS----------------
    elif slSum in percent6:
        print('range =', percent6)
        print(slSum, '= slSum')
        colorLen = 6
        count += 1
        sheet.cell(row=rowLen, column=17).value = colorList.cell(row=colorLen, column=2).value
        sheet.cell(row=rowLen, column=18).value = colorLen
        print(colorList.cell(row=colorLen, column=1).value)
    elif slSum in percent7:
        print('range =', percent7)
        print(slSum, '= slSum')
        colorLen = 7
        count += 1
        sheet.cell(row=rowLen, column=17).value = colorList.cell(row=colorLen, column=2).value
        sheet.cell(row=rowLen, column=18).value = colorLen
        print(colorList.cell(row=colorLen, column=1).value)
    elif slSum in percent8:
        print('range =', percent8)
        print(slSum, '= slSum')
        colorLen = 8
        count += 1
        sheet.cell(row=rowLen, column=17).value = colorList.cell(row=colorLen, column=2).value
        sheet.cell(row=rowLen, column=18).value = colorLen
        print(colorList.cell(row=colorLen, column=1).value)
    elif slSum in percent9:
        print('range =', percent9)
        print(slSum, '= slSum')
        colorLen = 9
        count += 1
        sheet.cell(row=rowLen, column=17).value = colorList.cell(row=colorLen, column=2).value
        sheet.cell(row=rowLen, column=18).value = colorLen

        print(colorList.cell(row=colorLen, column=1).value)
    elif slSum in percent10:
        print('range =', percent10)
        print(slSum, '= slSum')
        colorLen = 10
        count += 1
        sheet.cell(row=rowLen, column=17).value = colorList.cell(row=colorLen, column=2).value
        sheet.cell(row=rowLen, column=18).value = colorLen
        print(colorList.cell(row=colorLen, column=1).value)
    elif slSum >= 36 + 1:
        colorLen = 10
        count += 1
        sheet.cell(row=rowLen, column=17).value = colorList.cell(row=colorLen, column=2).value
        sheet.cell(row=rowLen, column=18).value = colorLen
        print(colorList.cell(row=colorLen, column=1).value)
    rowLen += 1

    print('count =', count)

url.save("main1.xlsx")
# ---------------------ЦИКЛ ПРОЦЕНТ ЗАКОНЧЕН------------------------
# </color>

# ----------------------------СОЗДАНИЕ КАРТЫ---------------------
m = folium.Map(location=[55.564638, 37.579508], zoom_start=15)  # выставляем центр обзора

figure_2 = folium.FeatureGroup(name="color markers").add_to(m)  # Создаем группу для определенной карты
# группы по процентам
figure_01 = folium.FeatureGroup(name='0-20 (%)').add_to(
    m)  # Создаем группу для определенного процента
figure_02 = folium.FeatureGroup(name='20-40 (%)').add_to(
    m)  # Создаем группу для определенного процента
figure_03 = folium.FeatureGroup(name='40-60 (%)').add_to(
    m)  # Создаем группу для оопределенного процента
figure_04 = folium.FeatureGroup(name='60-80 (%)').add_to(
    m)  # Создаем группу для определенного процента
figure_05 = folium.FeatureGroup(name='80-100 (%)').add_to(
    m)  # Создаем группу для определенного процента
figure_06 = folium.FeatureGroup(name='100-120 (%)').add_to(m)  # Создаем группу для определенной карты
figure_07 = folium.FeatureGroup(name='120-140 (%)').add_to(
    m)  # Создаем группу для определенного процента
figure_08 = folium.FeatureGroup(name='140-160 (%)').add_to(
    m)  # Создаем группу для определенного процента
figure_09 = folium.FeatureGroup(name='160-180 (%)').add_to(
    m)  # Создаем группу для определенного процента
figure_10 = folium.FeatureGroup(name='180-200 (%)').add_to(
    m)  # Создаем группу для определенного процента

# -------------КАРТА СОЗДАНА, НО ПУСТАЯ------------------

elevation = 2  # указываем высоту круга
rowAdd = 2  # номер строки
columCoordinates = 14  # номер столбца
ccd = 13  # номер соседнего столбца
for i in range(len(sort_BS)):
    icon_red = folium.Icon(color="green")  # придаем цвет точке на 2 карте
    # -----------------------КРУГ ВОЗЛЕ МАРКЕРА --------------
    folium.CircleMarker(((float(sheet.cell(row=rowAdd, column=columCoordinates).value),
                          float(sheet.cell(row=rowAdd, column=columCoordinates + 1).value))),
                        tooltip=(base_station[sheet.cell(row=rowAdd, column=columCoordinates - 1).value],
                                 sheet.cell(row=rowAdd, column=17).value), radius=15,
                        fill_opacity=0.9, popup=str(elevation), fill_color=sheet.cell(row=rowAdd, column=17).value
                        ).add_to(
        figure_2)  # создаем маркер на 2 карте
    print(i)  # проверка
    print('табличные координаты:', sheet.cell(row=rowAdd, column=columCoordinates).value)  # проверка данных
    print('табличные координаты тип:', type(
        float(sheet.cell(row=rowAdd, column=columCoordinates).value)))  # проверка данных с которыми работает цикл
    print(sheet.cell(row=rowAdd, column=ccd).value)  # проверка данных с которыми работает цикл
    rowAdd += 1  # переход на след строку
rowAdd = 2
for i in range(len(sort_BS)):

    icon_red = folium.Icon(color="green")  # придаем цвет точке на 2 карте
    if sheet.cell(row=rowAdd,
                  column=18).value == 1:  # сверяем указатель(в данном случае 1) в таблице и присваиваем к нужной группе
        folium.Marker(((float(sheet.cell(row=rowAdd, column=columCoordinates).value),
                        float(sheet.cell(row=rowAdd, column=columCoordinates + 1).value))),
                      tooltip=(base_station[sheet.cell(row=rowAdd, column=columCoordinates - 1).value],
                               sheet.cell(row=rowAdd, column=17).value, "0-20 (%)"), radius=15,
                      fill_opacity=0.9, popup=str(elevation), fill_color=sheet.cell(row=rowAdd, column=17).value
                      ).add_to(
            figure_01)  # указываем данные: координаты, название маркера
    # дальше все идет похожим методом
    elif sheet.cell(row=rowAdd, column=18).value == 2:
        folium.Marker(((float(sheet.cell(row=rowAdd, column=columCoordinates).value),
                        float(sheet.cell(row=rowAdd, column=columCoordinates + 1).value))),
                      tooltip=(base_station[sheet.cell(row=rowAdd, column=columCoordinates - 1).value],
                               sheet.cell(row=rowAdd, column=17).value, "20-40 (%)"), radius=15,
                      fill_opacity=0.9, popup=str(elevation), fill_color=sheet.cell(row=rowAdd, column=17).value
                      ).add_to(
            figure_02)
    elif sheet.cell(row=rowAdd, column=18).value == 3:
        folium.Marker(((float(sheet.cell(row=rowAdd, column=columCoordinates).value),
                        float(sheet.cell(row=rowAdd, column=columCoordinates + 1).value))),
                      tooltip=(base_station[sheet.cell(row=rowAdd, column=columCoordinates - 1).value],
                               sheet.cell(row=rowAdd, column=17).value, "40-60 (%)"), radius=15,
                      fill_opacity=0.9, popup=str(elevation), fill_color=sheet.cell(row=rowAdd, column=17).value
                      ).add_to(
            figure_03)
    elif sheet.cell(row=rowAdd, column=18).value == 4:
        folium.Marker(((float(sheet.cell(row=rowAdd, column=columCoordinates).value),
                        float(sheet.cell(row=rowAdd, column=columCoordinates + 1).value))),
                      tooltip=(base_station[sheet.cell(row=rowAdd, column=columCoordinates - 1).value],
                               sheet.cell(row=rowAdd, column=17).value, "60-80 (%)"), radius=15,
                      fill_opacity=0.9, popup=str(elevation), fill_color=sheet.cell(row=rowAdd, column=17).value
                      ).add_to(
            figure_04)
    elif sheet.cell(row=rowAdd, column=18).value == 5:
        folium.Marker(((float(sheet.cell(row=rowAdd, column=columCoordinates).value),
                        float(sheet.cell(row=rowAdd, column=columCoordinates + 1).value))),
                      tooltip=(base_station[sheet.cell(row=rowAdd, column=columCoordinates - 1).value],
                               sheet.cell(row=rowAdd, column=17).value, "80-100 (%)"), radius=15,
                      fill_opacity=0.9, popup=str(elevation), fill_color=sheet.cell(row=rowAdd, column=17).value
                      ).add_to(
            figure_05)
    elif sheet.cell(row=rowAdd, column=18).value == 6:
        folium.Marker(((float(sheet.cell(row=rowAdd, column=columCoordinates).value),
                        float(sheet.cell(row=rowAdd, column=columCoordinates + 1).value))),
                      tooltip=(base_station[sheet.cell(row=rowAdd, column=columCoordinates - 1).value],
                               sheet.cell(row=rowAdd, column=17).value, "100-120 (%)"), radius=15,
                      fill_opacity=0.9, popup=str(elevation), fill_color=sheet.cell(row=rowAdd, column=17).value
                      ).add_to(
            figure_06)
    elif sheet.cell(row=rowAdd, column=18).value == 7:
        folium.Marker(((float(sheet.cell(row=rowAdd, column=columCoordinates).value),
                        float(sheet.cell(row=rowAdd, column=columCoordinates + 1).value))),
                      tooltip=(base_station[sheet.cell(row=rowAdd, column=columCoordinates - 1).value],
                               sheet.cell(row=rowAdd, column=17).value, "120-140 (%)"), radius=15,
                      fill_opacity=0.9, popup=str(elevation), fill_color=sheet.cell(row=rowAdd, column=17).value
                      ).add_to(
            figure_07)
    elif sheet.cell(row=rowAdd, column=18).value == 8:
        folium.Marker(((float(sheet.cell(row=rowAdd, column=columCoordinates).value),
                        float(sheet.cell(row=rowAdd, column=columCoordinates + 1).value))),
                      tooltip=(base_station[sheet.cell(row=rowAdd, column=columCoordinates - 1).value],
                               sheet.cell(row=rowAdd, column=17).value, "140-160 (%)"), radius=15,
                      fill_opacity=0.9, popup=str(elevation), fill_color=sheet.cell(row=rowAdd, column=17).value
                      ).add_to(
            figure_08)
    elif sheet.cell(row=rowAdd, column=18).value == 9:
        folium.Marker(((float(sheet.cell(row=rowAdd, column=columCoordinates).value),
                        float(sheet.cell(row=rowAdd, column=columCoordinates + 1).value))),
                      tooltip=(base_station[sheet.cell(row=rowAdd, column=columCoordinates - 1).value],
                               sheet.cell(row=rowAdd, column=17).value, "160-180 (%)"), radius=15,
                      fill_opacity=0.9, popup=str(elevation), fill_color=sheet.cell(row=rowAdd, column=17).value
                      ).add_to(
            figure_09)
    elif sheet.cell(row=rowAdd, column=18).value == 10:
        folium.Marker(((float(sheet.cell(row=rowAdd, column=columCoordinates).value),
                        float(sheet.cell(row=rowAdd, column=columCoordinates + 1).value))),
                      tooltip=(base_station[sheet.cell(row=rowAdd, column=columCoordinates - 1).value],
                               sheet.cell(row=rowAdd, column=17).value, "180-200 (%)"), radius=15,
                      fill_opacity=0.9, popup=str(elevation), fill_color=sheet.cell(row=rowAdd, column=17).value
                      ).add_to(
            figure_10)
    # проверка работы цикла
    rowAdd += 1  # переход на след строку

url.save("main1.xlsx")  # сохраняем файл

folium.LayerControl(collapsed=True).add_to(m)  # Подключаем отображение списка групп
folium.LayerControl(collapsed=False).add_to(m)  # Подключаем отображение групп

m.save('Folium_1.html')  # сохраняем HTML файл
colorFile.close()
url.close()
